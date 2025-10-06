import csv
import io
import os
from datetime import date
from typing import Any, Dict, Iterable, List, Optional

import psycopg
from fastapi import FastAPI, File, HTTPException, UploadFile
from fastapi.responses import JSONResponse, StreamingResponse
from openpyxl import load_workbook
from pydantic import BaseModel, Field
from psycopg.rows import dict_row

SERVICE_NAME = os.getenv("SERVICE_NAME", "mask")
DSN = os.getenv("POSTGRES_DSN")

if not DSN:
    raise RuntimeError("POSTGRES_DSN is required for mask service")

app = FastAPI(title=SERVICE_NAME)


class Citation(BaseModel):
    norm_unit_id: Optional[str] = None
    chunk_id: Optional[str] = None


class KnowledgeAtomBase(BaseModel):
    title: str
    statement: str
    type: str
    bloom: str
    granularity: str
    status: str = "draft"
    version: Optional[str] = None
    prerequisites: List[str] = Field(default_factory=list)
    misconceptions: List[str] = Field(default_factory=list)
    citations: List[Citation] = Field(default_factory=list)


class KnowledgeAtomCreate(KnowledgeAtomBase):
    id: str


class KnowledgeAtomUpdate(BaseModel):
    title: Optional[str] = None
    statement: Optional[str] = None
    type: Optional[str] = None
    bloom: Optional[str] = None
    granularity: Optional[str] = None
    status: Optional[str] = None
    version: Optional[str] = None
    prerequisites: Optional[List[str]] = None
    misconceptions: Optional[List[str]] = None
    citations: Optional[List[Citation]] = None


class KnowledgeAtomOut(KnowledgeAtomBase):
    id: str
    created_at: Optional[str] = None
    updated_at: Optional[str] = None


class ContextProfileBase(BaseModel):
    title: str
    description: Optional[str] = None
    type: str
    bloom: str
    granularity: str
    status: str = "draft"
    version: Optional[str] = None
    prerequisites: List[str] = Field(default_factory=list)
    misconceptions: List[str] = Field(default_factory=list)
    citations: List[Citation] = Field(default_factory=list)
    atom_ids: List[str] = Field(default_factory=list)


class ContextProfileCreate(ContextProfileBase):
    id: str


class ContextProfileUpdate(BaseModel):
    title: Optional[str] = None
    description: Optional[str] = None
    type: Optional[str] = None
    bloom: Optional[str] = None
    granularity: Optional[str] = None
    status: Optional[str] = None
    version: Optional[str] = None
    prerequisites: Optional[List[str]] = None
    misconceptions: Optional[List[str]] = None
    citations: Optional[List[Citation]] = None
    atom_ids: Optional[List[str]] = None


class ContextProfileOut(ContextProfileBase):
    id: str
    created_at: Optional[str] = None
    updated_at: Optional[str] = None


class MCQImportResult(BaseModel):
    imported: int
    rejected_rows: List[int] = Field(default_factory=list)


@app.get("/health")
def health() -> Dict[str, str]:
    return {"status": "ok", "service": SERVICE_NAME}


def _conn():
    return psycopg.connect(DSN, row_factory=dict_row)


def _iso(value: Optional[date]) -> Optional[str]:
    return value.isoformat() if value else None


def _fetch_atom_citations(cur, atom_ids: Iterable[str]) -> Dict[str, List[Citation]]:
    atom_ids = list(atom_ids)
    mapping: Dict[str, List[Citation]] = {atom_id: [] for atom_id in atom_ids}
    if not atom_ids:
        return mapping
    cur.execute(
        """
        SELECT atom_id, norm_unit_id::text AS norm_unit_id, chunk_id::text AS chunk_id
        FROM atom_citation
        WHERE atom_id = ANY(%s)
        """,
        (atom_ids,),
    )
    for row in cur.fetchall():
        mapping.setdefault(row["atom_id"], []).append(
            Citation(norm_unit_id=row["norm_unit_id"], chunk_id=row["chunk_id"])
        )
    return mapping


def _fetch_context_citations(cur, context_ids: Iterable[str]) -> Dict[str, List[Citation]]:
    context_ids = list(context_ids)
    mapping: Dict[str, List[Citation]] = {ctx_id: [] for ctx_id in context_ids}
    if not context_ids:
        return mapping
    cur.execute(
        """
        SELECT context_id, norm_unit_id::text AS norm_unit_id, chunk_id::text AS chunk_id
        FROM context_citation
        WHERE context_id = ANY(%s)
        """,
        (context_ids,),
    )
    for row in cur.fetchall():
        mapping.setdefault(row["context_id"], []).append(
            Citation(norm_unit_id=row["norm_unit_id"], chunk_id=row["chunk_id"])
        )
    return mapping


def _atom_row_to_out(row: Dict[str, Any], citations: List[Citation]) -> KnowledgeAtomOut:
    return KnowledgeAtomOut(
        id=row["id"],
        title=row["title"],
        statement=row["statement"],
        type=row["type"],
        bloom=row["bloom"],
        granularity=row["granularity"],
        status=row["status"],
        version=row.get("version"),
        prerequisites=row.get("prerequisites") or [],
        misconceptions=row.get("misconceptions") or [],
        citations=citations,
        created_at=_iso(row.get("created_at")),
        updated_at=_iso(row.get("updated_at")),
    )


def _context_row_to_out(row: Dict[str, Any], citations: List[Citation]) -> ContextProfileOut:
    return ContextProfileOut(
        id=row["id"],
        title=row["title"],
        description=row.get("description"),
        type=row["type"],
        bloom=row["bloom"],
        granularity=row["granularity"],
        status=row["status"],
        version=row.get("version"),
        prerequisites=row.get("prerequisites") or [],
        misconceptions=row.get("misconceptions") or [],
        atom_ids=row.get("atom_ids") or [],
        citations=citations,
        created_at=_iso(row.get("created_at")),
        updated_at=_iso(row.get("updated_at")),
    )


def _replace_atom_citations(conn, atom_id: str, citations: Iterable[Citation]) -> None:
    with conn.cursor() as cur:
        cur.execute("DELETE FROM atom_citation WHERE atom_id = %s", (atom_id,))
        for item in citations:
            if not item.norm_unit_id and not item.chunk_id:
                continue
            cur.execute(
                """
                INSERT INTO atom_citation (atom_id, norm_unit_id, chunk_id)
                VALUES (%s, %s::uuid, %s::uuid)
                """,
                (atom_id, item.norm_unit_id, item.chunk_id),
            )


def _replace_context_citations(conn, context_id: str, citations: Iterable[Citation]) -> None:
    with conn.cursor() as cur:
        cur.execute("DELETE FROM context_citation WHERE context_id = %s", (context_id,))
        for item in citations:
            if not item.norm_unit_id and not item.chunk_id:
                continue
            cur.execute(
                """
                INSERT INTO context_citation (context_id, norm_unit_id, chunk_id)
                VALUES (%s, %s::uuid, %s::uuid)
                """,
                (context_id, item.norm_unit_id, item.chunk_id),
            )


@app.get("/atoms")
def list_atoms(limit: int = 100, offset: int = 0):
    with _conn() as conn, conn.cursor() as cur:
        cur.execute(
            """
            SELECT id, title, statement, type, bloom, granularity, status, version,
                   prerequisites, misconceptions, created_at, updated_at
            FROM knowledge_atom
            ORDER BY updated_at DESC
            LIMIT %s OFFSET %s
            """,
            (limit, offset),
        )
        rows = cur.fetchall()
        citations_map = _fetch_atom_citations(cur, [row["id"] for row in rows])
    items = [_atom_row_to_out(row, citations_map.get(row["id"], [])) for row in rows]
    return {"count": len(items), "items": [item.model_dump() for item in items]}


@app.post("/atoms", status_code=201)
def create_atom(payload: KnowledgeAtomCreate):
    with _conn() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                INSERT INTO knowledge_atom
                    (id, title, statement, type, bloom, granularity, status, version,
                     prerequisites, misconceptions, created_at, updated_at)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, now(), now())
                RETURNING id, title, statement, type, bloom, granularity, status, version,
                          prerequisites, misconceptions, created_at, updated_at
                """,
                (
                    payload.id,
                    payload.title,
                    payload.statement,
                    payload.type,
                    payload.bloom,
                    payload.granularity,
                    payload.status,
                    payload.version,
                    payload.prerequisites,
                    payload.misconceptions,
                ),
            )
            row = cur.fetchone()
        _replace_atom_citations(conn, payload.id, payload.citations)
        conn.commit()
    item = _atom_row_to_out(row, payload.citations)
    return JSONResponse(item.model_dump(), status_code=201)


@app.patch("/atoms/{atom_id}")
def update_atom(atom_id: str, payload: KnowledgeAtomUpdate):
    updates = payload.model_dump(exclude_unset=True)
    if not updates:
        raise HTTPException(status_code=400, detail="Empty payload")
    citations = updates.pop("citations", None)
    sets = []
    values: List[Any] = []
    for key, column in [
        ("title", "title"),
        ("statement", "statement"),
        ("type", "type"),
        ("bloom", "bloom"),
        ("granularity", "granularity"),
        ("status", "status"),
        ("version", "version"),
        ("prerequisites", "prerequisites"),
        ("misconceptions", "misconceptions"),
    ]:
        if key in updates:
            sets.append(f"{column} = %s")
            values.append(updates[key])
    if sets:
        sets.append("updated_at = now()")
        values.append(atom_id)
        with _conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    f"UPDATE knowledge_atom SET {', '.join(sets)} WHERE id = %s RETURNING id",
                    tuple(values),
                )
                row = cur.fetchone()
                if row is None:
                    raise HTTPException(status_code=404, detail="Atom not found")
            if citations is not None:
                _replace_atom_citations(conn, atom_id, citations)
            conn.commit()
    else:
        if citations is None:
            raise HTTPException(status_code=400, detail="Nothing to update")
        with _conn() as conn:
            _replace_atom_citations(conn, atom_id, citations)
            conn.commit()
    return get_atom(atom_id)


@app.get("/atoms/{atom_id}")
def get_atom(atom_id: str):
    with _conn() as conn, conn.cursor() as cur:
        cur.execute(
            """
            SELECT id, title, statement, type, bloom, granularity, status, version,
                   prerequisites, misconceptions, created_at, updated_at
            FROM knowledge_atom
            WHERE id = %s
            """,
            (atom_id,),
        )
        row = cur.fetchone()
        if row is None:
            raise HTTPException(status_code=404, detail="Atom not found")
        citations_map = _fetch_atom_citations(cur, [atom_id])
    item = _atom_row_to_out(row, citations_map.get(atom_id, []))
    return item.model_dump()


@app.delete("/atoms/{atom_id}", status_code=204)
def delete_atom(atom_id: str):
    with _conn() as conn, conn.cursor() as cur:
        cur.execute("DELETE FROM knowledge_atom WHERE id = %s", (atom_id,))
        if cur.rowcount == 0:
            raise HTTPException(status_code=404, detail="Atom not found")
        conn.commit()
    return JSONResponse(status_code=204, content=None)


@app.get("/contexts")
def list_contexts(limit: int = 100, offset: int = 0):
    with _conn() as conn, conn.cursor() as cur:
        cur.execute(
            """
            SELECT id, title, description, type, bloom, granularity, status, version,
                   prerequisites, misconceptions, atom_ids, created_at, updated_at
            FROM context_profile
            ORDER BY updated_at DESC
            LIMIT %s OFFSET %s
            """,
            (limit, offset),
        )
        rows = cur.fetchall()
        citations_map = _fetch_context_citations(cur, [row["id"] for row in rows])
    items = [_context_row_to_out(row, citations_map.get(row["id"], [])) for row in rows]
    return {"count": len(items), "items": [item.model_dump() for item in items]}


@app.post("/contexts", status_code=201)
def create_context(payload: ContextProfileCreate):
    with _conn() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                INSERT INTO context_profile
                    (id, title, description, type, bloom, granularity, status, version,
                     prerequisites, misconceptions, atom_ids, created_at, updated_at)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, now(), now())
                RETURNING id, title, description, type, bloom, granularity, status, version,
                          prerequisites, misconceptions, atom_ids, created_at, updated_at
                """,
                (
                    payload.id,
                    payload.title,
                    payload.description,
                    payload.type,
                    payload.bloom,
                    payload.granularity,
                    payload.status,
                    payload.version,
                    payload.prerequisites,
                    payload.misconceptions,
                    payload.atom_ids,
                ),
            )
            row = cur.fetchone()
        _replace_context_citations(conn, payload.id, payload.citations)
        conn.commit()
    item = _context_row_to_out(row, payload.citations)
    return JSONResponse(item.model_dump(), status_code=201)


@app.patch("/contexts/{context_id}")
def update_context(context_id: str, payload: ContextProfileUpdate):
    updates = payload.model_dump(exclude_unset=True)
    if not updates:
        raise HTTPException(status_code=400, detail="Empty payload")
    citations = updates.pop("citations", None)
    sets = []
    values: List[Any] = []
    for key, column in [
        ("title", "title"),
        ("description", "description"),
        ("type", "type"),
        ("bloom", "bloom"),
        ("granularity", "granularity"),
        ("status", "status"),
        ("version", "version"),
        ("prerequisites", "prerequisites"),
        ("misconceptions", "misconceptions"),
        ("atom_ids", "atom_ids"),
    ]:
        if key in updates:
            sets.append(f"{column} = %s")
            values.append(updates[key])
    if sets:
        sets.append("updated_at = now()")
        values.append(context_id)
        with _conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    f"UPDATE context_profile SET {', '.join(sets)} WHERE id = %s RETURNING id",
                    tuple(values),
                )
                row = cur.fetchone()
                if row is None:
                    raise HTTPException(status_code=404, detail="Context not found")
            if citations is not None:
                _replace_context_citations(conn, context_id, citations)
            conn.commit()
    else:
        if citations is None:
            raise HTTPException(status_code=400, detail="Nothing to update")
        with _conn() as conn:
            _replace_context_citations(conn, context_id, citations)
            conn.commit()
    return get_context(context_id)


@app.get("/contexts/{context_id}")
def get_context(context_id: str):
    with _conn() as conn, conn.cursor() as cur:
        cur.execute(
            """
            SELECT id, title, description, type, bloom, granularity, status, version,
                   prerequisites, misconceptions, atom_ids, created_at, updated_at
            FROM context_profile
            WHERE id = %s
            """,
            (context_id,),
        )
        row = cur.fetchone()
        if row is None:
            raise HTTPException(status_code=404, detail="Context not found")
        citations_map = _fetch_context_citations(cur, [context_id])
    item = _context_row_to_out(row, citations_map.get(context_id, []))
    return item.model_dump()


@app.delete("/contexts/{context_id}", status_code=204)
def delete_context(context_id: str):
    with _conn() as conn, conn.cursor() as cur:
        cur.execute("DELETE FROM context_profile WHERE id = %s", (context_id,))
        if cur.rowcount == 0:
            raise HTTPException(status_code=404, detail="Context not found")
        conn.commit()
    return JSONResponse(status_code=204, content=None)


def _parse_date(value: Any) -> Optional[date]:
    if value in (None, "", "-"):
        return None
    if isinstance(value, date):
        return value
    try:
        return date.fromisoformat(str(value))
    except ValueError as exc:  # pragma: no cover
        raise HTTPException(status_code=400, detail=f"Invalid date: {value}") from exc


def _load_question_choices(cur) -> Dict[str, List[Dict[str, Any]]]:
    cur.execute(
        """
        SELECT question_id::text AS question_id, text, is_correct, position
        FROM choice
        ORDER BY question_id, position
        """
    )
    mapping: Dict[str, List[Dict[str, Any]]] = {}
    for row in cur.fetchall():
        mapping.setdefault(row["question_id"], []).append(row)
    return mapping


@app.post("/questions/import_xlsx", response_model=MCQImportResult)
async def import_questions(file: UploadFile = File(...)):
    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files are supported")
    raw = await file.read()
    try:
        workbook = load_workbook(io.BytesIO(raw))
    except Exception as exc:  # pragma: no cover - delegated to runtime
        raise HTTPException(status_code=400, detail="Failed to read XLSX") from exc
    sheet = workbook.active
    header = [str(cell.value).strip().lower() if cell.value else "" for cell in next(sheet.iter_rows(max_row=1))]
    expected = ["prompt", "option1", "option2", "option3", "option4", "correct", "norm_unit_id"]
    if any(col not in header for col in expected):
        raise HTTPException(status_code=400, detail=f"Missing required columns: {expected}")
    indices = {name: header.index(name) for name in header if name}
    imported = 0
    rejected: List[int] = []
    with _conn() as conn:
        for idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
            values = [cell.value for cell in row]
            prompt = str(values[indices["prompt"]]).strip() if values[indices["prompt"]] else ""
            norm_unit_raw = values[indices["norm_unit_id"]]
            norm_unit_id = str(norm_unit_raw).strip() if norm_unit_raw else None
            if not prompt or not norm_unit_id:
                rejected.append(idx)
                continue
            choices = []
            for i in range(1, 5):
                val = values[indices.get(f"option{i}", -1)] if indices.get(f"option{i}") is not None else None
                if val is None or str(val).strip() == "":
                    rejected.append(idx)
                    choices = []
                    break
                choices.append(str(val).strip())
            if not choices:
                continue
            correct_raw = values[indices["correct"]]
            try:
                correct_idx = int(correct_raw) - 1
            except (TypeError, ValueError):
                correct_text = str(correct_raw).strip() if correct_raw else ""
                correct_idx = next((i for i, choice in enumerate(choices) if choice == correct_text), -1)
            if correct_idx < 0 or correct_idx >= len(choices):
                rejected.append(idx)
                continue
            chunk_raw = values[indices.get("chunk_id", -1)] if indices.get("chunk_id") is not None else None
            chunk_id = str(chunk_raw).strip() if chunk_raw else None
            valid_from = _parse_date(values[indices.get("valid_from", -1)]) if indices.get("valid_from") is not None else None
            valid_to = _parse_date(values[indices.get("valid_to", -1)]) if indices.get("valid_to") is not None else None
            with conn.cursor() as cur:
                cur.execute(
                    """
                    INSERT INTO question (prompt, norm_unit_id, chunk_id, valid_from, valid_to, meta, created_at, updated_at)
                    VALUES (%s, %s::uuid, %s::uuid, %s, %s, %s, now(), now())
                    RETURNING id::text
                    """,
                    (prompt, norm_unit_id, chunk_id, valid_from, valid_to, None),
                )
                question_id = cur.fetchone()["id"]
                for pos, text in enumerate(choices):
                    cur.execute(
                        """
                        INSERT INTO choice (question_id, text, is_correct, position)
                        VALUES (%s::uuid, %s, %s, %s)
                        """,
                        (question_id, text, pos == correct_idx, pos + 1),
                    )
            imported += 1
        conn.commit()
    return MCQImportResult(imported=imported, rejected_rows=rejected)


def _compose_norm_ref(row: Dict[str, Any]) -> Optional[str]:
    segments: List[str] = []
    if row.get("law"):
        segments.append(row["law"])
    if row.get("article"):
        segments.append(f"статья {row['article']}")
    if row.get("part"):
        segments.append(f"часть {row['part']}")
    if row.get("point"):
        segments.append(f"пункт {row['point']}")
    if row.get("subpoint"):
        segments.append(f"подпункт {row['subpoint']}")
    return ", ".join(segments) if segments else None


@app.get("/cards/export/anki")
def export_cards():
    output = io.StringIO()
    writer = csv.writer(output, delimiter="\t")
    with _conn() as conn, conn.cursor() as cur:
        cur.execute(
            """
            SELECT q.id::text AS id, q.prompt, q.norm_unit_id::text AS norm_unit_id, q.chunk_id::text AS chunk_id,
                   q.valid_from, q.valid_to,
                   n.law, n.article, n.part, n.point, n.subpoint
            FROM question AS q
            LEFT JOIN norm_unit AS n ON q.norm_unit_id = n.id
            ORDER BY q.created_at DESC
            """
        )
        questions = cur.fetchall()
        choices_map = _load_question_choices(cur)
    writer.writerow(["Front", "Back", "NormReference", "ValidFrom", "ValidTo"])
    for question in questions:
        choices = choices_map.get(question["id"], [])
        correct = [item["text"] for item in choices if item["is_correct"]]
        distractors = [item["text"] for item in choices if not item["is_correct"]]
        back_parts = []
        if correct:
            back_parts.append("Верный ответ: " + "; ".join(correct))
        if distractors:
            back_parts.append("Другие варианты: " + "; ".join(distractors))
        norm_ref = _compose_norm_ref(question)
        writer.writerow(
            [
                question["prompt"],
                "<br>".join(back_parts) if back_parts else "",
                norm_ref or "",
                _iso(question.get("valid_from")),
                _iso(question.get("valid_to")),
            ]
        )
    output.seek(0)
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode("utf-8")),
        media_type="text/tab-separated-values",
        headers={
            "Content-Disposition": 'attachment; filename="anki_cards.tsv"',
        },
    )
